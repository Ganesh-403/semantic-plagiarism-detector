import inspect
import io
import os
from datetime import datetime, timezone

import openpyxl
import pandas as pd
import pytest

from src.utils.bulk_export import export_incidents_xlsx_stream
from src.utils.excel_export import (
    build_similarity_workbook,
    export_similarity_matrix_to_excel,
    export_similarity_matrix_to_temp_file,
    generate_csv_matrix_stream,
    generate_tsv_matrix_stream,
)
from src.utils.export_sanitizer import sanitize_spreadsheet_value


def test_generate_csv_matrix_stream():
    # Setup test DataFrame
    data = {
        "DocA.txt": [1.0, 0.85, 0.12],
        "DocB.txt": [0.85, 1.0, 0.45],
        "DocC.txt": [0.12, 0.45, 1.0],
    }
    df = pd.DataFrame(data, index=["DocA.txt", "DocB.txt", "DocC.txt"])

    # Test 1: Return type is a Generator
    stream = generate_csv_matrix_stream(df)
    assert inspect.isgenerator(stream)

    # Test 2: Verify chunk output
    chunks = list(stream)
    assert len(chunks) == len(df) + 1  # 1 header row + 3 data rows

    # Verify header line
    assert chunks[0].strip() == "Document,DocA.txt,DocB.txt,DocC.txt"

    # Verify data lines
    assert chunks[1].strip() == "DocA.txt,1.0,0.85,0.12"
    assert chunks[2].strip() == "DocB.txt,0.85,1.0,0.45"
    assert chunks[3].strip() == "DocC.txt,0.12,0.45,1.0"

    # Test 3: Verify complete CSV reconstruction matches Expected CSV output
    full_csv = "".join(chunks)
    reconstructed_df = pd.read_csv(io.StringIO(full_csv), index_col=0)
    pd.testing.assert_frame_equal(df, reconstructed_df, check_names=False)


def test_generate_tsv_matrix_stream():
    # Setup test DataFrame
    data = {
        "DocA.txt": [1.0, 0.85, 0.12],
        "DocB.txt": [0.85, 1.0, 0.45],
        "DocC.txt": [0.12, 0.45, 1.0],
    }
    df = pd.DataFrame(data, index=["DocA.txt", "DocB.txt", "DocC.txt"])

    # Test 1: Return type is a Generator
    stream = generate_tsv_matrix_stream(df)
    assert inspect.isgenerator(stream)

    # Test 2: Verify chunk output
    chunks = list(stream)
    assert len(chunks) == len(df) + 1  # 1 header row + 3 data rows

    # Verify header line with tabs
    assert chunks[0].strip() == "Document\tDocA.txt\tDocB.txt\tDocC.txt"

    # Verify data lines with tabs
    assert chunks[1].strip() == "DocA.txt\t1.0\t0.85\t0.12"
    assert chunks[2].strip() == "DocB.txt\t0.85\t1.0\t0.45"
    assert chunks[3].strip() == "DocC.txt\t0.12\t0.45\t1.0"

    # Test 3: Verify complete TSV reconstruction matches Expected TSV output
    full_tsv = "".join(chunks)
    reconstructed_df = pd.read_csv(io.StringIO(full_tsv), sep="\t", index_col=0)
    pd.testing.assert_frame_equal(df, reconstructed_df, check_names=False)



def test_build_similarity_workbook_metadata_properties():
    """Verify build_similarity_workbook populates document title, creator, and created timestamp (#3438)."""
    df = pd.DataFrame({"Doc1.txt": [1.0]}, index=["Doc1.txt"])
    before = datetime.now(timezone.utc)
    wb = build_similarity_workbook(df)
    after = datetime.now(timezone.utc)

    assert wb.properties.title == "Semantic Plagiarism Similarity Report"
    assert wb.properties.creator == "Semantic Plagiarism Detector"
    assert wb.properties.created is not None
    assert isinstance(wb.properties.created, datetime)
    assert before <= wb.properties.created <= after


def test_build_similarity_workbook_custom_color_thresholds():
    df = pd.DataFrame(
        {"a.txt": [1.0, 0.4], "b.txt": [0.4, 1.0]},
        index=["a.txt", "b.txt"],
    )
    wb = build_similarity_workbook(
        df,
        low_threshold=0.1,
        mid_threshold=0.4,
        high_threshold=0.9,
    )
    rules = list(wb.active.conditional_formatting._cf_rules.values())
    rule = rules[0][0]
    assert float(rule.colorScale.cfvo[0].val) == 0.1
    assert float(rule.colorScale.cfvo[1].val) == 0.4
    assert float(rule.colorScale.cfvo[2].val) == 0.9


def test_export_similarity_matrix_to_excel_persists_metadata():
    """Verify export_similarity_matrix_to_excel persists metadata in the saved XLSX file (#3438)."""
    df = pd.DataFrame(
        {"DocA.txt": [1.0, 0.8], "DocB.txt": [0.8, 1.0]},
        index=["DocA.txt", "DocB.txt"],
    )
    xlsx_bytes = export_similarity_matrix_to_excel(df)
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 0

    # Load back with openpyxl to inspect file properties
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.properties.title == "Semantic Plagiarism Similarity Report"
    assert wb.properties.creator == "Semantic Plagiarism Detector"
    assert wb.properties.created is not None


def test_export_incidents_xlsx_stream_persists_metadata():
    """Verify export_incidents_xlsx_stream sets title, creator, and created metadata (#3438)."""
    incidents = [
        {
            "incident_id": "INC-001",
            "document_a": "Essay1.docx",
            "document_b": "Essay2.docx",
            "similarity_score": 0.88,
            "severity_rank": "High",
            "review_status": "Pending",
            "date_flagged": "2026-08-25",
        }
    ]
    xlsx_bytes = export_incidents_xlsx_stream(incidents)
    assert isinstance(xlsx_bytes, bytes)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.properties.title == "Semantic Plagiarism Similarity Report"
    assert wb.properties.creator == "Semantic Plagiarism Detector"
    assert wb.properties.created is not None


def test_build_similarity_workbook_nan_handling():
    """Verify build_similarity_workbook handles NaN/None values gracefully (#3437)."""
    data = {
        "DocA.txt": [1.0, float("nan"), 0.3],
        "DocB.txt": [0.8, 1.0, None],
        "DocC.txt": [0.2, 0.7, 1.0],
    }
    df = pd.DataFrame(data, index=["DocA.txt", "DocB.txt", "DocC.txt"])

    wb = build_similarity_workbook(df)
    ws = wb.active

    # Cell with NaN should contain "-"
    assert ws.cell(row=3, column=2).value == "-"
    # Cell with None should contain "-"
    assert ws.cell(row=4, column=3).value == "-"
    # Normal values should still be floats
    assert ws.cell(row=2, column=2).value == 1.0
    assert ws.cell(row=2, column=3).value == 0.8


def test_export_similarity_matrix_to_excel_with_nan():
    """Verify export_similarity_matrix_to_excel produces valid XLSX with NaN values (#3437)."""
    data = {
        "DocA.txt": [1.0, float("nan")],
        "DocB.txt": [0.5, 1.0],
    }
    df = pd.DataFrame(data, index=["DocA.txt", "DocB.txt"])

    xlsx_bytes = export_similarity_matrix_to_excel(df)
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 0

    # Load back and verify NaN cell is "-"
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.cell(row=2, column=3).value == "-"


def test_build_similarity_workbook_flagged_pairs_standard():
    """Verify that build_similarity_workbook creates the 'Flagged Pairs' sheet with correct severity classification in standard mode."""
    data = {
        "DocA.txt": [1.0, 0.97, 0.88, 0.50],
        "DocB.txt": [0.97, 1.0, 0.72, 0.40],
        "DocC.txt": [0.88, 0.72, 1.0, 0.30],
        "DocD.txt": [0.50, 0.40, 0.30, 1.0],
    }
    df = pd.DataFrame(data, index=["DocA.txt", "DocB.txt", "DocC.txt", "DocD.txt"])

    # Build workbook with 0.70 threshold
    wb = build_similarity_workbook(df, threshold=0.70)
    
    assert "Flagged Pairs" in wb.sheetnames
    ws = wb["Flagged Pairs"]
    
    # Check headers
    assert ws.cell(row=1, column=1).value == "Document A"
    assert ws.cell(row=1, column=2).value == "Document B"
    assert ws.cell(row=1, column=3).value == "Similarity Score"
    assert ws.cell(row=1, column=4).value == "Severity"
    
    # Flagged pairs should be:
    # 1. DocA.txt - DocB.txt (0.97) -> Critical
    # 2. DocA.txt - DocC.txt (0.88) -> High
    # 3. DocB.txt - DocC.txt (0.72) -> Moderate
    # (Only 3 rows total in addition to header)
    assert ws.max_row == 4
    
    # Row 2 (DocA-DocB)
    assert ws.cell(row=2, column=1).value == "DocA.txt"
    assert ws.cell(row=2, column=2).value == "DocB.txt"
    assert ws.cell(row=2, column=3).value == 0.97
    assert ws.cell(row=2, column=4).value == "Critical"
    
    # Row 3 (DocA-DocC)
    assert ws.cell(row=3, column=1).value == "DocA.txt"
    assert ws.cell(row=3, column=2).value == "DocC.txt"
    assert ws.cell(row=3, column=3).value == 0.88
    assert ws.cell(row=3, column=4).value == "High"
    
    # Row 4 (DocB-DocC)
    assert ws.cell(row=4, column=1).value == "DocB.txt"
    assert ws.cell(row=4, column=2).value == "DocC.txt"
    assert ws.cell(row=4, column=3).value == 0.72
    assert ws.cell(row=4, column=4).value == "Moderate"


def test_build_similarity_workbook_flagged_pairs_write_only():
    """Verify that build_similarity_workbook creates the 'Flagged Pairs' sheet in write-only mode."""
    data = {
        "DocA.txt": [1.0, 0.96, 0.75],
        "DocB.txt": [0.96, 1.0, 0.40],
        "DocC.txt": [0.75, 0.40, 1.0],
    }
    df = pd.DataFrame(data, index=["DocA.txt", "DocB.txt", "DocC.txt"])

    wb = build_similarity_workbook(df, threshold=0.70, write_only=True)
    
    assert "Flagged Pairs" in wb.sheetnames
    
    # Save write-only workbook to bytes to parse it back with read-only openpyxl
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    saved_wb = openpyxl.load_workbook(out, read_only=True)
    assert "Flagged Pairs" in saved_wb.sheetnames
    ws = saved_wb["Flagged Pairs"]
    
    # Get rows
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 3  # Header + 2 flagged pairs (DocA-DocB: 0.96, DocA-DocC: 0.75)
    
    assert rows[0] == ("Document A", "Document B", "Similarity Score", "Severity")
    assert rows[1] == ("DocA.txt", "DocB.txt", 0.96, "Critical")
    assert rows[2] == ("DocA.txt", "DocC.txt", 0.75, "Moderate")


