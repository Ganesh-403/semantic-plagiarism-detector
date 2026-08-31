"""
src/utils/excel_export.py
-------------------------
Utility for exporting similarity matrices into styled Excel (.xlsx) workbooks
with conditional formatting matching the application's heatmap logic.
Supports both in-memory generation and managed temporary disk-file creation with automatic exit cleanup.
Also provides streaming CSV generation for memory-efficient exports of large datasets.
"""

import atexit
import csv
import io
import os
import re
import tempfile
from datetime import datetime, timezone
from typing import Generator

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.utils.export_sanitizer import (
    FORMULA_TRIGGER_PREFIXES,
    sanitize_spreadsheet_value,
)

# Excel rejects these characters outright in a worksheet title, and caps the
# title at 31 characters. Titles can originate from a course or assignment
# name, so they are not trustworthy input. The set is the one established in
# #3673 plus the backslash, which Excel rejects as well but which that pass
# missed.
_INVALID_SHEET_TITLE_CHARS = re.compile(r"[\[\]\*\?:/\\.]")

#: Excel's hard limit on worksheet title length.
MAX_SHEET_TITLE_LENGTH = 31

#: Used when sanitization consumes the whole title.
DEFAULT_SHEET_TITLE = "Sheet"

#: Default worksheet title for an exported similarity matrix.
DEFAULT_WORKSHEET_TITLE = "Similarity Matrix"


def sanitize_sheet_title(title) -> str:
    """Coerce a worksheet title into something Excel will accept.

    Excel worksheet titles cannot exceed 31 characters and cannot contain
    ``[``, ``]``, ``*``, ``?``, ``:``, ``/``, ``\\`` or ``.``. openpyxl raises
    when handed a title that breaks either rule, which would abort the export.

    Args:
        title: The desired title. Coerced to ``str``, so a non-string label
            (e.g. an integer assignment ID) does not raise.

    Returns:
        A title that satisfies both rules. Falls back to
        :data:`DEFAULT_SHEET_TITLE` when sanitization leaves nothing usable,
        because openpyxl also rejects an empty title.
    """
    sanitized = _INVALID_SHEET_TITLE_CHARS.sub("", str(title))
    # Excel additionally rejects a title that is only whitespace, and trims
    # surrounding whitespace itself; do it here so the length cap is applied
    # to what actually lands in the file.
    sanitized = sanitized.strip()[:MAX_SHEET_TITLE_LENGTH].strip()

    return sanitized or DEFAULT_SHEET_TITLE


def _create_managed_temp_file(suffix: str = ".xlsx", prefix: str = "temp_") -> str:
    """Helper to create a temporary file that is automatically deleted on exit."""
    fd, temp_path = tempfile.mkstemp(suffix=suffix, prefix=prefix)
    os.close(fd)

    def _cleanup():
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass

    atexit.register(_cleanup)
    return temp_path


def _truncate_title(title, max_length: int = 60) -> str:
    """
    Truncate a title to max_length characters, appending '...' if truncated.

    Args:
        title: The title to truncate. Coerced to str, so a non-string
            DataFrame index (e.g. an integer document ID) does not raise.
        max_length: Maximum length before truncation (default: 60)

    Returns:
        Truncated title with '...' suffix if original was longer
    """
    title = str(title)
    if len(title) <= max_length:
        return title
    return title[: max_length - 3] + "..."


def build_similarity_workbook(
    df: pd.DataFrame,
    threshold: float = 0.59,
    write_only: bool = False,
    sheet_title: str = DEFAULT_WORKSHEET_TITLE,
    low_threshold: float = 0.0,
    mid_threshold: float = 0.59,
    high_threshold: float = 1.0,
) -> Workbook:
    """Helper function that builds and styles the openpyxl Workbook.

    Args:
        df: Similarity matrix DataFrame with document labels as index and columns.
        threshold: Score threshold for conditional formatting color scale.
        write_only: If True, uses openpyxl write_only mode with ws.append() for
            memory-efficient streaming of large matrices. Defaults to False.
        sheet_title: Worksheet title. Passed through
            :func:`sanitize_sheet_title`, so a caller may hand in an untrusted
            label (a course or assignment name) without the export aborting.
        low_threshold: Low breakpoint for the 3-color scale.
        mid_threshold: Mid breakpoint for the 3-color scale.
        high_threshold: High breakpoint for the 3-color scale.

    Returns:
        Workbook: Configured openpyxl Workbook instance.
    """
    safe_sheet_title = sanitize_sheet_title(sheet_title)

    # Older callers pass ``threshold`` as the yellow midpoint.
    if threshold != 0.59 and mid_threshold == 0.59:
        mid_threshold = threshold

    if write_only:
        wb = Workbook(write_only=True)
        wb.properties.title = "Semantic Plagiarism Similarity Report"
        wb.properties.creator = "Semantic Plagiarism Detector"
        wb.properties.created = datetime.now(timezone.utc)

        ws = wb.create_sheet(title=safe_sheet_title)

        header_fill = PatternFill(
            start_color="1F2937", end_color="1F2937", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")

        # Write header row
        header_row = []
        c0 = WriteOnlyCell(ws, value="Document")
        c0.fill = header_fill
        c0.font = header_font
        c0.alignment = header_align
        header_row.append(c0)

        for col_name in df.columns:
            truncated_name = _truncate_title(col_name)
            cell = WriteOnlyCell(
                ws, value=sanitize_spreadsheet_value(truncated_name)
            )
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            if len(str(col_name)) > 60:
                cell.comment = Comment(
                    sanitize_spreadsheet_value(str(col_name)), "Excel Export"
                )
            header_row.append(cell)
        ws.append(header_row)

        # Write data rows
        for index_label, row in df.iterrows():
            row_cells = []
            truncated_label = _truncate_title(index_label)
            label_cell = WriteOnlyCell(
                ws, value=sanitize_spreadsheet_value(truncated_label)
            )
            label_cell.fill = header_fill
            label_cell.font = header_font
            if len(str(index_label)) > 60:
                label_cell.comment = Comment(
                    sanitize_spreadsheet_value(str(index_label)), "Excel Export"
                )
            row_cells.append(label_cell)

            for val in row:
                val_cell = WriteOnlyCell(ws, value=float(val))
                val_cell.number_format = "0.0%"
                val_cell.alignment = Alignment(horizontal="right")
                row_cells.append(val_cell)

            ws.append(row_cells)

        # Apply Conditional Formatting (3-Color Scale)
        max_row = len(df) + 1
        max_col = len(df.columns) + 1

        if max_row > 1 and max_col > 1:
            start_cell = "B2"
            end_col_letter = get_column_letter(max_col)
            end_cell = f"{end_col_letter}{max_row}"
            matrix_range = f"{start_cell}:{end_cell}"

            color_scale = ColorScaleRule(
                start_type="num",
                start_value=low_threshold,
                start_color="FFFFFF",  # White (low)
                mid_type="num",
                mid_value=mid_threshold,
                mid_color="FEF08A",  # Yellow (mid)
                end_type="num",
                end_value=high_threshold,
                end_color="EF4444",  # Red (high)
            )
            ws.conditional_formatting.add(matrix_range, color_scale)

        # Auto-adjust column widths
        max_index_len = max([len(str(idx)) for idx in df.index] + [len("Document")])
        ws.column_dimensions[get_column_letter(1)].width = max(
            min(max_index_len, 60) + 3, 12
        )

        for col_idx, col_name in enumerate(df.columns, start=2):
            col_len = min(len(str(col_name)), 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(
                col_len + 3, 12
            )

        # Create Flagged Pairs worksheet (write-only mode)
        flagged_ws = wb.create_sheet(title="Flagged Pairs")

        # Write header row for flagged sheet
        flagged_header_row = []
        for col_name in ["Document A", "Document B", "Similarity Score", "Severity"]:
            cell = WriteOnlyCell(flagged_ws, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            flagged_header_row.append(cell)
        flagged_ws.append(flagged_header_row)

        # Populate flagged pairs (upper triangle only)
        docs = list(df.index)
        for i in range(len(docs)):
            for j in range(i + 1, len(docs)):
                val = df.iloc[i, j]
                if pd.notna(val) and float(val) >= mid_threshold:
                    score = float(val)
                    severity = "Moderate"
                    if score >= 0.85:
                        severity = "High"
                    if score >= 0.95:
                        severity = "Critical"

                    row_data = [
                        sanitize_spreadsheet_value(str(docs[i])),
                        sanitize_spreadsheet_value(str(docs[j])),
                        score,
                        severity,
                    ]
                    row_cells = []
                    for c_idx, item in enumerate(row_data):
                        cell = WriteOnlyCell(flagged_ws, value=item)
                        if c_idx == 2:
                            cell.number_format = "0.0%"
                            cell.alignment = Alignment(horizontal="right")
                        row_cells.append(cell)
                    flagged_ws.append(row_cells)

        # Set column widths for flagged sheet
        flagged_ws.column_dimensions["A"].width = 25
        flagged_ws.column_dimensions["B"].width = 25
        flagged_ws.column_dimensions["C"].width = 18
        flagged_ws.column_dimensions["D"].width = 15

        return wb

    # Default write_only=False (in-memory DOM)
    wb = Workbook()
    wb.properties.title = "Semantic Plagiarism Similarity Report"
    wb.properties.creator = "Semantic Plagiarism Detector"
    wb.properties.created = datetime.now(timezone.utc)

    ws = wb.active
    ws.title = safe_sheet_title

    # Write headers and index labels with truncated titles, preserving full names in comments.
    # Labels originate from uploaded filenames, so they are sanitized before
    # being written to prevent formula injection in the exported workbook.
    ws.cell(row=1, column=1, value="Document")
    for col_idx, col_name in enumerate(df.columns, start=2):
        truncated_name = _truncate_title(col_name)
        cell = ws.cell(
            row=1, column=col_idx, value=sanitize_spreadsheet_value(truncated_name)
        )
        # Add full title as comment if truncated
        if len(str(col_name)) > 60:
            cell.comment = Comment(
                sanitize_spreadsheet_value(str(col_name)), "Excel Export"
            )

    for row_idx, (index_label, row) in enumerate(df.iterrows(), start=2):
        truncated_label = _truncate_title(index_label)
        cell = ws.cell(
            row=row_idx, column=1, value=sanitize_spreadsheet_value(truncated_label)
        )
        # Add full title as comment if truncated
        if len(str(index_label)) > 60:
            cell.comment = Comment(
                sanitize_spreadsheet_value(str(index_label)), "Excel Export"
            )

        for col_idx, val in enumerate(row, start=2):
            if pd.isna(val):
                cell = ws.cell(row=row_idx, column=col_idx, value="-")
                cell.alignment = Alignment(horizontal="right")
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=float(val))
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="right")

    # Header styling
    header_fill = PatternFill(
        start_color="1F2937", end_color="1F2937", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font

    # Apply Conditional Formatting (3-Color Scale)
    max_row = len(df) + 1
    max_col = len(df.columns) + 1

    if max_row > 1 and max_col > 1:
        start_cell = "B2"
        end_col_letter = ws.cell(row=max_row, column=max_col).column_letter
        end_cell = f"{end_col_letter}{max_row}"
        matrix_range = f"{start_cell}:{end_cell}"

        color_scale = ColorScaleRule(
            start_type="num",
            start_value=low_threshold,
            start_color="FFFFFF",  # White (low)
            mid_type="num",
            mid_value=mid_threshold,
            mid_color="FEF08A",  # Yellow (mid)
            end_type="num",
            end_value=high_threshold,
            end_color="EF4444",  # Red (high)
        )
        ws.conditional_formatting.add(matrix_range, color_scale)

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Create Flagged Pairs worksheet
    flagged_ws = wb.create_sheet(title=sanitize_sheet_title("Flagged Pairs"))

    # Write headers and style them
    flagged_ws.cell(row=1, column=1, value="Document A")
    flagged_ws.cell(row=1, column=2, value="Document B")
    flagged_ws.cell(row=1, column=3, value="Similarity Score")
    flagged_ws.cell(row=1, column=4, value="Severity")

    for cell in flagged_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Populate flagged pairs (upper triangle only)
    docs = list(df.index)
    current_row = 2
    for i in range(len(docs)):
        for j in range(i + 1, len(docs)):
            val = df.iloc[i, j]
            if pd.notna(val) and float(val) >= mid_threshold:
                score = float(val)
                severity = "Moderate"
                if score >= 0.85:
                    severity = "High"
                if score >= 0.95:
                    severity = "Critical"

                flagged_ws.cell(
                    row=current_row,
                    column=1,
                    value=sanitize_spreadsheet_value(str(docs[i])),
                )
                flagged_ws.cell(
                    row=current_row,
                    column=2,
                    value=sanitize_spreadsheet_value(str(docs[j])),
                )

                score_cell = flagged_ws.cell(row=current_row, column=3, value=score)
                score_cell.number_format = "0.0%"
                score_cell.alignment = Alignment(horizontal="right")

                flagged_ws.cell(row=current_row, column=4, value=severity)
                current_row += 1

    # Set column widths for flagged sheet
    flagged_ws.column_dimensions["A"].width = 25
    flagged_ws.column_dimensions["B"].width = 25
    flagged_ws.column_dimensions["C"].width = 18
    flagged_ws.column_dimensions["D"].width = 15

    return wb


def export_similarity_matrix_to_excel(
    df: pd.DataFrame,
    threshold: float = 0.59,
    write_only: bool = False,
    sheet_title: str = DEFAULT_WORKSHEET_TITLE,
) -> bytes:
    """Exports a similarity matrix DataFrame into an in-memory Excel file (.xlsx) with formatting."""
    wb = build_similarity_workbook(
        df,
        threshold=threshold,
        write_only=write_only,
        sheet_title=sheet_title,
        mid_threshold=threshold,
    )
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_similarity_matrix_to_temp_file(
    df: pd.DataFrame,
    threshold: float = 0.59,
    write_only: bool = False,
    sheet_title: str = DEFAULT_WORKSHEET_TITLE,
) -> str:
    """
    Exports the similarity matrix to a temporary .xlsx file on disk.
    The created file is automatically registered for cleanup on application exit via atexit.

    Returns:
        str: Absolute path to the created temporary Excel file.
    """
    wb = build_similarity_workbook(
        df,
        threshold=threshold,
        write_only=write_only,
        sheet_title=sheet_title,
        mid_threshold=threshold,
    )
    temp_path = _create_managed_temp_file(suffix=".xlsx", prefix="similarity_matrix_")
    wb.save(temp_path)
    return temp_path


def generate_csv_matrix_stream(matrix_df: pd.DataFrame) -> Generator[str, None, None]:
    """
    Yields CSV formatted lines line-by-line from a similarity matrix DataFrame.

    Memory-efficient generator for exporting large result sets (>10,000 document pairs)
    without materializing the entire formatted output string or Excel workbook in memory.

    Args:
        matrix_df (pd.DataFrame): Similarity matrix DataFrame with document labels as index and columns.

    Yields:
        str: CSV formatted string row (including newline character).
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # Yield header row. Column labels are sanitized for the same reason as in
    # build_similarity_workbook(): a CSV opened in Excel evaluates formulas too,
    # so the streaming route must not be a way around the protection.
    header = ["Document"] + [sanitize_spreadsheet_value(c) for c in matrix_df.columns]
    writer.writerow(header)
    yield buffer.getvalue()
    buffer.seek(0)
    buffer.truncate(0)

    # Yield data rows line by line
    for index, row in matrix_df.iterrows():
        writer.writerow(
            [sanitize_spreadsheet_value(index)]
            + [sanitize_spreadsheet_value(v) for v in row.tolist()]
        )
        yield buffer.getvalue()
        buffer.seek(0)
        buffer.truncate(0)


def generate_tsv_matrix_stream(matrix_df: pd.DataFrame) -> Generator[str, None, None]:
    """
    Yields TSV formatted lines line-by-line from a similarity matrix DataFrame.

    Memory-efficient generator for exporting large result sets (>10,000 document pairs)
    using tab-delimited formatting for R and Pandas workflows.

    Args:
        matrix_df (pd.DataFrame): Similarity matrix DataFrame with document labels as index and columns.

    Yields:
        str: TSV formatted string row (including newline character).
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter="\t")

    # Yield header row
    header = ["Document"] + [sanitize_spreadsheet_value(c) for c in matrix_df.columns]
    writer.writerow(header)
    yield buffer.getvalue()
    buffer.seek(0)
    buffer.truncate(0)

    # Yield data rows line by line
    for index, row in matrix_df.iterrows():
        writer.writerow(
            [sanitize_spreadsheet_value(index)]
            + [sanitize_spreadsheet_value(v) for v in row.tolist()]
        )
        yield buffer.getvalue()
        buffer.seek(0)
        buffer.truncate(0)

